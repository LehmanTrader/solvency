#!/usr/bin/env python3
"""Export every bench run to one Excel workbook for audit and revisiting.

    python3 bench/export-excel.py [out.xlsx]

Sheets:
  runs      one row per run (summary.json): model, harness, access, basis,
            protocol, trials, pass rate, $/task, $/solved, status flags
  attempts  one row per attempt (results.jsonl): run, task, trial, pass,
            detail, cost, token usage, latency, reply head
Runs marked INVALID or SUPERSEDED are included and flagged — the audit trail
is the point — but never confused with countable runs.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
RESULTS = ROOT / "results"
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Desktop" / "solvency-bench-logs.xlsx"

RUN_COLS = [
    "run_id", "status", "model", "harness", "harness_version", "access",
    "cost_basis", "protocol", "trials", "countable_attempts", "pass_rate",
    "cost_per_task_usd", "cost_per_solved_usd", "run_date",
]
ATT_COLS = [
    "run_id", "task_id", "trial", "pass", "infra", "detail", "cost_usd",
    "input_tokens", "cache_read", "cache_write", "output_tokens", "ms", "at", "reply_head",
]

wb = Workbook()
runs_ws = wb.active
runs_ws.title = "runs"
att_ws = wb.create_sheet("attempts")
bold = Font(bold=True)

runs_ws.append(RUN_COLS)
att_ws.append(ATT_COLS)
for ws in (runs_ws, att_ws):
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"

n_runs = n_att = 0
for run_dir in sorted(RESULTS.iterdir()):
    if not run_dir.is_dir():
        continue
    summary_path = run_dir / "summary.json"
    status = "OK"
    if (run_dir / "INVALID.json").exists():
        status = "INVALID"
    elif (run_dir / "SUPERSEDED.json").exists():
        status = "SUPERSEDED"
    s = {}
    if summary_path.exists():
        s = json.loads(summary_path.read_text())
    elif status == "OK":
        status = "INCOMPLETE"
    harness = s.get("harness") or {}
    runs_ws.append([
        run_dir.name, status, s.get("model"),
        harness.get("name") if isinstance(harness, dict) else harness,
        (str(harness.get("version")).split("\n")[0] if isinstance(harness, dict) and harness.get("version") else None),
        s.get("access"), s.get("costBasis"), s.get("protocol"), s.get("trials"),
        s.get("countableAttempts") or s.get("attempts_countable"),
        s.get("passRate"), s.get("costPerTaskUsd"), s.get("costPerSolvedUsd"),
        s.get("runDate") or s.get("run_date"),
    ])
    n_runs += 1
    results_path = run_dir / "results.jsonl"
    if results_path.exists():
        for line in results_path.read_text().splitlines():
            if not line.strip():
                continue
            try:
                a = json.loads(line)
            except json.JSONDecodeError:
                continue
            u = a.get("usage") or {}
            att_ws.append([
                run_dir.name, a.get("taskId"), a.get("trial"),
                a.get("pass"), a.get("infra"), a.get("detail"), a.get("costUsd"),
                u.get("input"), u.get("cacheRead"), u.get("cacheWrite"), u.get("output"),
                a.get("ms"), a.get("at"),
                (a.get("replyHead") or "")[:200],
            ])
            n_att += 1

for ws, widths in ((runs_ws, [34, 11, 26, 12, 22, 30, 30, 20, 7, 10, 9, 14, 15, 12]),
                   (att_ws, [34, 18, 6, 6, 6, 44, 10, 9, 10, 11, 9, 8, 22, 60])):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"wrote {OUT}: {n_runs} runs, {n_att} attempts")
